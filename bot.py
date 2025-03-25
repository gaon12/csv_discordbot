import csv
import discord
from discord.ext import commands
import asyncio
import logging
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import openpyxl
import xlrd
import os
import json
import shutil

# 로깅 설정
logging.basicConfig(level=logging.INFO)

def read_token_from_file(file_path):
    try:
        with open(file_path, 'r') as f:
            token = f.read().strip()
        return token
    except Exception as e:
        logging.error(f"Token file read error: {e}")
        return None

TOKEN = read_token_from_file('token.txt')

def read_csv_or_tsv(file_path, delimiter=','):
    try:
        with open(file_path, mode='r', encoding='utf-8') as f:
            reader = csv.reader(f, delimiter=delimiter)
            next(reader)  # 첫 번째 행(헤더) 건너뛰기
            return {row[0]: row[1] for row in reader}
    except Exception as e:
        logging.error(f"Error reading CSV/TSV file: {e}")
        return {}

def read_xls(file_path):
    try:
        workbook = xlrd.open_workbook(file_path)
        sheet = workbook.sheet_by_index(0)
        return {sheet.cell_value(row, 0): sheet.cell_value(row, 1) for row in range(1, sheet.nrows)}
    except Exception as e:
        logging.error(f"Error reading XLS file: {e}")
        return {}

def read_xlsx(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active
        return {sheet.cell(row=row, column=1).value: sheet.cell(row=row, column=2).value for row in range(2, sheet.max_row + 1)}
    except Exception as e:
        logging.error(f"Error reading XLSX file: {e}")
        return {}

def read_json(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        logging.error(f"Error reading JSON file: {e}")
        return {}

def save_to_custom_file(data, original_file_path):
    try:
        file_extension = os.path.splitext(original_file_path)[1]
        custom_file_path = original_file_path.replace(file_extension, '_custom' + file_extension)

        if file_extension in ['.csv', '.tsv']:
            with open(custom_file_path, mode='w', encoding='utf-8', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['Question', 'Answer'])  # 헤더 작성
                for question, answer in data.items():
                    writer.writerow([question, answer])
        elif file_extension == '.xlsx':
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(['Question', 'Answer'])
            for question, answer in data.items():
                sheet.append([question, answer])
            workbook.save(custom_file_path)
        elif file_extension == '.xls':
            workbook = xlrd.open_workbook(original_file_path)
            sheet = workbook.sheet_by_index(0)
            # 저장할 새로운 워크북을 만들어 기존 데이터를 복사하고 추가
            new_workbook = xlrd.open_workbook(custom_file_path)
            new_sheet = new_workbook.sheet_by_index(0)
            new_sheet.write(0, 0, 'Question')
            new_sheet.write(0, 1, 'Answer')
            for question, answer in data.items():
                new_sheet.write(new_sheet.max_row, 0, question)
                new_sheet.write(new_sheet.max_row, 1, answer)
            new_workbook.save(custom_file_path)

        logging.info(f"Custom file saved at {custom_file_path}")
    except Exception as e:
        logging.error(f"Error saving to custom file: {e}")

async def load_qa_data():
    file_path = 'qa_data'  # 파일 이름(확장자 제외)
    if os.path.exists(f"{file_path}.tsv"):
        return read_csv_or_tsv(f"{file_path}.tsv", delimiter='\t')
    elif os.path.exists(f"{file_path}.xlsx"):
        return read_xlsx(f"{file_path}.xlsx")
    elif os.path.exists(f"{file_path}.xls"):
        return read_xls(f"{file_path}.xls")
    elif os.path.exists(f"{file_path}.csv"):
        return read_csv_or_tsv(f"{file_path}.csv")
    elif os.path.exists(f"{file_path}.json"):
        return read_json(f"{file_path}.json")
    else:
        logging.error("No suitable QA data file found.")
        return {}

class FileChangeHandler(FileSystemEventHandler):
    def __init__(self, bot, filename, callback):
        self.bot = bot
        self.filename = filename
        self.callback = callback

    def on_modified(self, event):
        if event.src_path == self.filename:
            asyncio.run_coroutine_threadsafe(self.callback(), self.bot.loop)

intents = discord.Intents.default()
intents.messages = True
intents.guilds = True
intents.message_content = True

bot = commands.Bot(command_prefix='!', intents=intents)

@bot.event
async def on_ready():
    global qa_data
    qa_data = await load_qa_data()
    print(f'{bot.user} has connected to Discord!')

@bot.event
async def on_message(message):
    if message.author == bot.user:
        return
    await bot.process_commands(message)

@bot.command()
async def cmdname(ctx, *args):
    question = ' '.join(args)
    if question in qa_data:
        await ctx.send(qa_data[question])
    else:
        await ctx.send('Sorry, no answer values were found for the question you entered.')

@bot.command()
async def help_cmd(ctx):
    await ctx.send("Usage: !cmdname [question]")

@bot.command()
async def addqa(ctx, question, answer):
    qa_data[question] = answer
    await ctx.send(f"QA pair added: {question} -> {answer}")
    logging.info(f"Added QA pair: {question} -> {answer}")
    
    # 원본 파일 경로 찾기
    original_file_path = 'qa_data.csv'  # 원본 파일 경로 (확장자 포함)
    save_to_custom_file(qa_data, original_file_path)

async def update_qa_data():
    global qa_data
    qa_data = await load_qa_data()
    logging.info("QA data updated.")

if TOKEN:
    event_handler = FileChangeHandler(bot, 'qa_data.csv', update_qa_data)
    observer = Observer()
    observer.schedule(event_handler, path='.', recursive=False)
    observer.start()
    try:
        bot.run(TOKEN)
    finally:
        observer.stop()
        observer.join()
else:
    logging.error("No valid token found.")
